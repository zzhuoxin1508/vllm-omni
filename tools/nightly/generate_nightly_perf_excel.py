#!/usr/bin/env python3
"""
Generate a nightly Excel performance report from JSON results.

"""

from __future__ import annotations

import argparse
import json
import logging
import os
from collections.abc import Iterable, Sequence
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

LOGGER = logging.getLogger(__name__)

GREY_BLOCK_FILL = PatternFill(start_color="D3D3D3", fill_type="solid")

# Diffusion sheet columns (Qwen-Image diffusion benchmark).
DIFFUSION_BENCHMARK_COLUMNS: tuple[str, ...] = (
    "duration",
    "completed_requests",
    "failed_requests",
    "throughput_qps",
    "latency_mean",
    "latency_median",
    "latency_p50",
    "latency_p99",
    "peak_memory_mb_max",
    "peak_memory_mb_mean",
    "peak_memory_mb_median",
    "slo_attainment_rate",
)

DIFFUSION_NUMERIC_FORMAT_COLUMNS: tuple[str, ...] = DIFFUSION_BENCHMARK_COLUMNS

DIFFUSION_SUMMARY_COLUMNS: tuple[str, ...] = (
    "date",
    "test_name",
    "model",
    "backend",
    "dataset",
    "task",
    "completed_requests",
    "failed_requests",
    "duration",
    "throughput_qps",
    "latency_mean",
    "latency_median",
    "latency_p50",
    "latency_p99",
    "peak_memory_mb_max",
    "peak_memory_mb_mean",
    "peak_memory_mb_median",
    "slo_attainment_rate",
    "commit_sha",
    "build_id",
    "build_url",
    "source_file",
)

# Benchmark metric columns: grey the latest row's cell when value changed vs previous date.
BENCHMARK_COLUMNS: tuple[str, ...] = (
    "num_prompts",
    "request_rate",
    "burstiness",
    "max_concurrency",
    "duration",
    "completed",
    "failed",
    "request_throughput",
    "output_throughput",
    "total_token_throughput",
    "mean_ttft_ms",
    "p99_ttft_ms",
    "mean_tpot_ms",
    "p99_tpot_ms",
    "mean_itl_ms",
    "p99_itl_ms",
    "mean_e2el_ms",
    "p99_e2el_ms",
    "mean_audio_rtf",
    "p99_audio_rtf",
    "mean_audio_duration_s",
    "p99_audio_duration_s",
)
# Columns that get float coercion and number format in Excel. Excludes request_rate ("inf" str)
# and max_concurrency (null); leave those as-is. If they become float in the future, they are
# still written correctly from JSON without coercion here.
NUMERIC_FORMAT_COLUMNS: tuple[str, ...] = tuple(
    c for c in BENCHMARK_COLUMNS if c not in ("request_rate", "max_concurrency")
)
DATASET_NAME_ALLOWED = ("random", "random-mm")

_COLUMNS_FILENAME = "nightly_perf_summary_columns.txt"
_RESULT_JSON_PREFIX = "result_test_"
_DIFFUSION_JSON_PREFIX = "diffusion_perf_"
DEFAULT_INPUT_DIR = os.getenv("DEFAULT_INPUT_DIR") if os.getenv("DEFAULT_INPUT_DIR") else "tests"
DEFAULT_OUTPUT_DIR = os.getenv("DEFAULT_OUTPUT_DIR") if os.getenv("DEFAULT_OUTPUT_DIR") else "tests"
DEFAULT_DIFFUSION_INPUT_DIR = os.getenv("DIFFUSION_BENCHMARK_DIR")
# Read omni/diffusion benchmarks from DEFAULT_INPUT_DIR, if DEFAULT_DIFFUSION_INPUT_DIR is not set.


def _omni_group_key(record: dict[str, Any]) -> tuple[Any, ...]:
    return (
        record.get("model_id") or "",
        record.get("test_name") or "",
        record.get("dataset_name") or "",
        record.get("max_concurrency") if record.get("max_concurrency") is not None else 0,
        record.get("num_prompts") if record.get("num_prompts") is not None else 0,
    )


def _diffusion_group_key(record: dict[str, Any]) -> tuple[Any, ...]:
    return (record.get("test_name") or "",)


def _load_summary_columns(script_dir: str) -> list[str]:
    """Load summary column names from a file next to this script; fallback to default if missing."""
    path = os.path.join(script_dir, _COLUMNS_FILENAME)
    default = [
        "date",
        "endpoint_type",
        "backend",
        "model_id",
        "tokenizer_id",
        "test_name",
        "dataset_name",
        "num_prompts",
        "request_rate",
        "burstiness",
        "max_concurrency",
        "duration",
        "completed",
        "failed",
        "request_throughput",
        "output_throughput",
        "total_token_throughput",
        "mean_ttft_ms",
        "p99_ttft_ms",
        "mean_tpot_ms",
        "p99_tpot_ms",
        "mean_itl_ms",
        "p99_itl_ms",
        "mean_e2el_ms",
        "p99_e2el_ms",
        "mean_audio_rtf",
        "p99_audio_rtf",
        "mean_audio_duration_s",
        "p99_audio_duration_s",
        "commit_sha",
        "build_id",
        "build_url",
        "source_file",
    ]
    if not os.path.isfile(path):
        return default
    columns: list[str] = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if s and not s.startswith("#"):
                columns.append(s)
    return columns if columns else default


def _ensure_omni_summary_columns(summary_columns: list[str]) -> list[str]:
    """Ensure omni summary contains required columns, even when a custom columns file exists."""
    required = ("test_name", "dataset_name", "source_file")
    existing = set(summary_columns)
    if all(c in existing for c in required):
        return summary_columns

    out = list(summary_columns)
    insert_after = "tokenizer_id"
    if "test_name" not in existing:
        if insert_after in existing:
            idx = out.index(insert_after) + 1
            out.insert(idx, "test_name")
        else:
            out.append("test_name")
    existing = set(out)
    if "dataset_name" not in existing:
        if "test_name" in out:
            idx = out.index("test_name") + 1
            out.insert(idx, "dataset_name")
        else:
            out.append("dataset_name")
    if "source_file" not in set(out):
        out.append("source_file")
    return out


def _vllm_omni_root() -> str:
    """Resolve vllm-omni repo root: directory that contains a 'tests' subdir (and usually 'tools')."""
    path = os.path.dirname(os.path.abspath(__file__))
    while path and path != os.path.dirname(path):
        if os.path.isdir(os.path.join(path, "tests")):
            return path
        path = os.path.dirname(path)
    return os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))


def _default_input_dir() -> str:
    """Default: vllm-omni root / DEFAULT_INPUT_DIR (where performance JSON files live)."""
    root = _vllm_omni_root()
    return os.path.join(root, DEFAULT_INPUT_DIR)


def _default_diffusion_input_dir(input_dir: str) -> str:
    """Default diffusion input dir: DIFFUSION_BENCHMARK_DIR if set, else fall back to omni input dir."""
    return DEFAULT_DIFFUSION_INPUT_DIR if DEFAULT_DIFFUSION_INPUT_DIR else input_dir


def _default_output_file() -> str:
    """Default: vllm-omni root / DEFAULT_OUTPUT_DIR / nightly_perf_<timestamp>.xlsx."""
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return os.path.join(_vllm_omni_root(), DEFAULT_OUTPUT_DIR, f"nightly_perf_{ts}.xlsx")


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Read performance JSON files from vllm-omni/tests/ and generate an Excel report."
    )
    parser.add_argument(
        "--input-dir",
        type=str,
        default=_default_input_dir(),
        help="Directory containing performance JSON files; default is <vllm-omni-root>/DEFAULT_INPUT_DIR.",
    )
    parser.add_argument(
        "--diffusion-input-dir",
        type=str,
        default=None,
        help=(
            "Directory containing diffusion_perf_*.json files; default is "
            "DIFFUSION_BENCHMARK_DIR, fallback to --input-dir."
        ),
    )
    parser.add_argument(
        "--output-file",
        type=str,
        default=_default_output_file(),
        help=(
            "Output path of the Excel report; default is "
            "<vllm-omni-root>/DEFAULT_OUTPUT_DIR/nightly_perf_<timestamp>.xlsx."
        ),
    )
    parser.add_argument(
        "--commit-sha",
        type=str,
        default=None,
        help="Optional commit SHA; defaults to environment variable BUILDKITE_COMMIT if unset.",
    )
    parser.add_argument(
        "--build-id",
        type=str,
        default=None,
        help="Optional build ID; defaults to environment variable BUILDKITE_BUILD_ID if unset.",
    )
    parser.add_argument(
        "--build-url",
        type=str,
        default=None,
        help="Optional build URL; defaults to environment variable BUILDKITE_BUILD_URL if unset.",
    )
    return parser.parse_args()


def _load_json_file(path: str) -> dict[str, Any] | None:
    """Safely load a single JSON file; return None and log a warning on failure."""
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as exc:
        LOGGER.warning("failed to load json '%s': %s", path, exc)
        return None

    if not isinstance(data, dict):
        LOGGER.warning("json root in '%s' is not an object, skip", path)
        return None

    return data


def _parse_from_filename(filename: str) -> dict[str, Any]:
    """Parse test-related metadata from a result JSON filename.

    Expected pattern (after prefix/suffix stripped):
    <test_name>_<dataset_name>_<max_concurrency>_<num_prompts>_<timestamp>
    """
    name, ext = os.path.splitext(filename)
    if ext != ".json" or not name.startswith(_RESULT_JSON_PREFIX):
        return {}

    core = name[len(_RESULT_JSON_PREFIX) :]
    parts = core.split("_")
    if len(parts) < 5:
        LOGGER.warning("filename '%s' does not match expected pattern, skip parsing test metadata", filename)
        return {}

    timestamp = parts[-1]
    num_prompts_str = parts[-2]
    max_concurrency_str = parts[-3]
    dataset_name = parts[-4]
    test_name = "_".join(parts[:-4]) if parts[:-4] else ""

    parsed: dict[str, Any] = {}

    if len(timestamp) >= 15:
        parsed["date"] = timestamp

    if dataset_name in DATASET_NAME_ALLOWED:
        parsed["dataset_name"] = dataset_name

    try:
        parsed["num_prompts"] = int(num_prompts_str)
    except (TypeError, ValueError):
        pass

    try:
        parsed["max_concurrency"] = int(max_concurrency_str)
    except (TypeError, ValueError):
        pass

    if test_name:
        parsed["test_name"] = test_name

    return parsed


def _iter_omni_json_records(input_dir: str) -> Iterable[dict[str, Any]]:
    """Iterate over result_test_*.json files and yield normalized omni records."""
    if not os.path.isdir(input_dir):
        LOGGER.warning("input dir '%s' does not exist or is not a directory", input_dir)
        return

    for entry in sorted(os.listdir(input_dir)):
        if not entry.endswith(".json"):
            continue
        if not entry.startswith(_RESULT_JSON_PREFIX):
            continue
        full_path = os.path.join(input_dir, entry)
        if not os.path.isfile(full_path):
            continue

        data = _load_json_file(full_path)
        if data is None:
            continue

        record: dict[str, Any] = dict(data)
        filename_meta = _parse_from_filename(os.path.basename(full_path))

        if "date" not in record or not record["date"]:
            if "date" in filename_meta:
                record["date"] = filename_meta["date"]
            else:
                record["date"] = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")

        if "num_prompts" not in record or record["num_prompts"] is None:
            if "num_prompts" in filename_meta:
                record["num_prompts"] = filename_meta["num_prompts"]

        if "max_concurrency" not in record or record["max_concurrency"] is None:
            if "max_concurrency" in filename_meta:
                record["max_concurrency"] = filename_meta["max_concurrency"]

        if "test_name" not in record or not record.get("test_name"):
            if "test_name" in filename_meta:
                record["test_name"] = filename_meta["test_name"]

        if "dataset_name" not in record or not record.get("dataset_name"):
            if "dataset_name" in filename_meta:
                record["dataset_name"] = filename_meta["dataset_name"]

        record["source_file"] = os.path.basename(full_path)
        yield record


def _parse_diffusion_from_filename(filename: str) -> dict[str, Any]:
    """Parse diffusion test_name/date from filename: diffusion_perf_<test_name>_<YYYYMMDD-HHMMSS>.json"""
    name, ext = os.path.splitext(filename)
    if ext != ".json" or not name.startswith(_DIFFUSION_JSON_PREFIX):
        return {}
    core = name[len(_DIFFUSION_JSON_PREFIX) :]
    parts = core.split("_")
    if len(parts) < 2:
        return {}
    timestamp = parts[-1]
    test_name = "_".join(parts[:-1]) if parts[:-1] else ""
    parsed: dict[str, Any] = {}
    if len(timestamp) >= 15:
        parsed["date"] = timestamp
    if test_name:
        parsed["test_name"] = test_name
    return parsed


def _iter_diffusion_json_records(input_dir: str) -> Iterable[dict[str, Any]]:
    """Iterate over diffusion_perf_*.json files and yield normalized diffusion records."""
    if not os.path.isdir(input_dir):
        LOGGER.warning("diffusion input dir '%s' does not exist or is not a directory", input_dir)
        return

    for entry in sorted(os.listdir(input_dir)):
        if not entry.endswith(".json"):
            continue
        if not entry.startswith(_DIFFUSION_JSON_PREFIX):
            continue
        full_path = os.path.join(input_dir, entry)
        if not os.path.isfile(full_path):
            continue

        data = _load_json_file(full_path)
        if data is None:
            continue

        record: dict[str, Any] = dict(data)
        filename_meta = _parse_diffusion_from_filename(os.path.basename(full_path))
        if "date" not in record or not record.get("date"):
            record["date"] = filename_meta.get("date") or datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        if "test_name" not in record or not record.get("test_name"):
            if "test_name" in filename_meta:
                record["test_name"] = filename_meta["test_name"]
        record["source_file"] = os.path.basename(full_path)
        yield record


def _collect_records(input_dir: str) -> list[dict[str, Any]]:
    return list(_iter_omni_json_records(input_dir))


def _collect_diffusion_records(diffusion_input_dir: str) -> list[dict[str, Any]]:
    return list(_iter_diffusion_json_records(diffusion_input_dir))


def _apply_build_metadata_to_latest_only(
    records: Sequence[dict[str, Any]],
    commit_sha: str | None,
    build_id: str | None,
    build_url: str | None,
) -> None:
    """Set commit_sha, build_id, build_url only on rows with the latest date.
    Other rows get None so that build info is not duplicated for older benchmark data.
    """
    if not records:
        return
    max_date = max((r.get("date") or "") for r in records)
    for r in records:
        if (r.get("date") or "") == max_date:
            r["commit_sha"] = commit_sha
            r["build_id"] = build_id
            r["build_url"] = build_url
        else:
            r["commit_sha"] = None
            r["build_id"] = None
            r["build_url"] = None


def _sort_records_for_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sort so that same test configuration is grouped, newest date first within each group."""
    by_date_desc = sorted(records, key=lambda r: (r.get("date") or ""), reverse=True)
    return sorted(
        by_date_desc,
        key=_omni_group_key,
    )


def _sort_diffusion_records_for_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_date_desc = sorted(records, key=lambda r: (r.get("date") or ""), reverse=True)
    return sorted(by_date_desc, key=_diffusion_group_key)


def _values_differ(a: Any, b: Any) -> bool:
    """Return True if two values are considered different; avoid direct float equality."""
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    if isinstance(a, float) and isinstance(b, float):
        if a != a and b != b:
            return False
        if a != a or b != b:
            return True
        return abs(a - b) > 1e-9
    return a != b


def _apply_change_highlight(
    ws,
    columns: Sequence[str],
    records: Sequence[dict[str, Any]],
    benchmark_columns: Sequence[str],
    group_key_fn,
) -> None:
    """Grey cells in the latest row of each group when a metric changed vs previous row in the same group."""
    if not records:
        return
    col_to_index = {c: i + 1 for i, c in enumerate(columns)}
    i = 0
    while i < len(records):
        group_key = group_key_fn(records[i])
        block_start = i
        while i < len(records) and group_key_fn(records[i]) == group_key:
            i += 1
        block_end = i
        newest_idx = block_start
        prev_idx = block_start + 1 if block_start + 1 < block_end else None
        if prev_idx is None:
            continue
        excel_row = newest_idx + 2
        for col in benchmark_columns:
            if col not in col_to_index:
                continue
            cur_val = records[newest_idx].get(col)
            prev_val = records[prev_idx].get(col)
            if _values_differ(cur_val, prev_val):
                ws.cell(row=excel_row, column=col_to_index[col]).fill = GREY_BLOCK_FILL


def _build_raw_columns(
    records: Sequence[dict[str, Any]],
    summary_columns: Sequence[str],
) -> list[str]:
    """Infer the column set for the raw sheet based on all records."""
    keys: set[str] = set()
    for record in records:
        keys.update(record.keys())
    # Ensure summary columns appear first; remaining columns sorted alphabetically.
    ordered_keys: list[str] = []
    for key in summary_columns:
        if key in keys:
            ordered_keys.append(key)
            keys.discard(key)
    ordered_keys.extend(sorted(keys))
    return ordered_keys


def _to_float_if_numeric(value: Any) -> Any:
    """Coerce to float when possible so Excel treats as number; avoid #### from narrow columns."""
    if value is None:
        return value
    if isinstance(value, (int, float)):
        return float(value) if isinstance(value, int) else value
    if isinstance(value, str):
        try:
            return float(value)
        except (ValueError, TypeError):
            return value
    return value


def _write_sheet(
    ws,
    columns: Sequence[str],
    rows: Iterable[dict[str, Any]],
    numeric_columns: Sequence[str] = (),
) -> None:
    """Write column names and row data into the given worksheet."""
    numeric_set = set(numeric_columns)
    ws.append(list(columns))
    for record in rows:
        row_values = []
        for col in columns:
            v = record.get(col)
            if col in numeric_set:
                v = _to_float_if_numeric(v)
            row_values.append(v)
        ws.append(row_values)


def _format_numeric_columns(
    ws,
    columns: Sequence[str],
    numeric_columns: Sequence[str],
    num_data_rows: int,
    width: int,
) -> None:
    """Set number format and column width for numeric columns so values display (no ####)."""
    numeric_set = set(numeric_columns)
    for c, col_name in enumerate(columns):
        if col_name not in numeric_set:
            continue
        col_letter = get_column_letter(c + 1)
        ws.column_dimensions[col_letter].width = width
        for r in range(2, 2 + num_data_rows):
            cell = ws.cell(row=r, column=c + 1)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"
            elif isinstance(cell.value, str):
                try:
                    cell.value = float(cell.value)
                    cell.number_format = "0.0000"
                except (ValueError, TypeError):
                    pass


def _set_column_width(ws, columns: Sequence[str], col_name: str, width: int) -> None:
    """Set a single column width by column name if present."""
    try:
        idx = list(columns).index(col_name)
    except ValueError:
        return
    ws.column_dimensions[get_column_letter(idx + 1)].width = width


def _apply_column_widths(ws, columns: Sequence[str], widths: dict[str, int]) -> None:
    for name, width in widths.items():
        _set_column_width(ws, columns, name, width)


_OMNI_SUMMARY_WIDTHS = {
    "endpoint_type": 10,
    "backend": 10,
    "model_id": 20,
    "tokenizer_id": 20,
    "test_name": 20,
    "dataset_name": 10,
}

_DIFFUSION_SUMMARY_WIDTHS = {
    "test_name": 30,
    "model": 15,
}


def _ensure_parent_dir(path: str) -> None:
    """Ensure that the parent directory of the output file exists."""
    parent = os.path.dirname(os.path.abspath(path))
    if not parent:
        return
    os.makedirs(parent, exist_ok=True)


def generate_excel_report(
    input_dir: str,
    diffusion_input_dir: str,
    output_file: str,
    commit_sha: str | None,
    build_id: str | None,
    build_url: str | None,
) -> None:
    """Main logic: load JSON records and generate an Excel report with multiple sheets."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    omni_summary_columns = _ensure_omni_summary_columns(_load_summary_columns(script_dir))

    omni_records = _collect_records(input_dir)
    diffusion_records = _collect_diffusion_records(diffusion_input_dir)

    if not omni_records:
        LOGGER.warning("no valid omni json records found under '%s'", input_dir)
    if not diffusion_records:
        LOGGER.warning("no valid diffusion json records found under '%s'", diffusion_input_dir)

    omni_sorted = _sort_records_for_summary(omni_records)
    diffusion_sorted = _sort_diffusion_records_for_summary(diffusion_records)

    _apply_build_metadata_to_latest_only(omni_sorted, commit_sha, build_id, build_url)
    _apply_build_metadata_to_latest_only(diffusion_sorted, commit_sha, build_id, build_url)

    wb = Workbook()
    ws_omni_summary = wb.active
    ws_omni_summary.title = "omni_summary"
    _write_sheet(ws_omni_summary, omni_summary_columns, omni_sorted, numeric_columns=NUMERIC_FORMAT_COLUMNS)
    _format_numeric_columns(
        ws_omni_summary,
        omni_summary_columns,
        NUMERIC_FORMAT_COLUMNS,
        len(omni_sorted),
        width=14,
    )
    _apply_column_widths(ws_omni_summary, omni_summary_columns, _OMNI_SUMMARY_WIDTHS)
    _apply_change_highlight(
        ws_omni_summary,
        omni_summary_columns,
        omni_sorted,
        BENCHMARK_COLUMNS,
        _omni_group_key,
    )

    ws_diff_summary = wb.create_sheet(title="diffusion_summary")
    _write_sheet(
        ws_diff_summary, DIFFUSION_SUMMARY_COLUMNS, diffusion_sorted, numeric_columns=DIFFUSION_NUMERIC_FORMAT_COLUMNS
    )
    _format_numeric_columns(
        ws_diff_summary,
        DIFFUSION_SUMMARY_COLUMNS,
        DIFFUSION_NUMERIC_FORMAT_COLUMNS,
        len(diffusion_sorted),
        width=16,
    )
    _apply_column_widths(ws_diff_summary, DIFFUSION_SUMMARY_COLUMNS, _DIFFUSION_SUMMARY_WIDTHS)
    _apply_change_highlight(
        ws_diff_summary,
        DIFFUSION_SUMMARY_COLUMNS,
        diffusion_sorted,
        DIFFUSION_BENCHMARK_COLUMNS,
        _diffusion_group_key,
    )

    if omni_sorted:
        omni_raw_columns = _build_raw_columns(omni_sorted, omni_summary_columns)
        ws_omni_raw = wb.create_sheet(title="omni_raw")
        _write_sheet(ws_omni_raw, omni_raw_columns, omni_sorted)

    if diffusion_sorted:
        diffusion_raw_columns = _build_raw_columns(diffusion_sorted, DIFFUSION_SUMMARY_COLUMNS)
        ws_diff_raw = wb.create_sheet(title="diffusion_raw")
        _write_sheet(ws_diff_raw, diffusion_raw_columns, diffusion_sorted)

    _ensure_parent_dir(output_file)
    wb.save(output_file)
    LOGGER.info("excel report saved to '%s'", output_file)


def main() -> None:
    """Command-line entrypoint."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    args = parse_args()

    commit_sha = args.commit_sha or os.getenv("BUILDKITE_COMMIT")
    build_id = args.build_id or os.getenv("BUILDKITE_BUILD_ID")
    build_url = args.build_url or os.getenv("BUILDKITE_BUILD_URL")
    diffusion_input_dir = args.diffusion_input_dir or _default_diffusion_input_dir(args.input_dir)

    generate_excel_report(
        input_dir=args.input_dir,
        diffusion_input_dir=diffusion_input_dir,
        output_file=args.output_file,
        commit_sha=commit_sha,
        build_id=build_id,
        build_url=build_url,
    )


if __name__ == "__main__":
    main()
