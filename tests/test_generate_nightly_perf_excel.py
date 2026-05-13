import importlib.util
import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

pytestmark = [pytest.mark.core_model, pytest.mark.cpu]


def _load_excel_module():
    repo_root = Path(__file__).resolve().parents[1]
    module_path = repo_root / "tools" / "nightly" / "generate_nightly_perf_excel.py"
    spec = importlib.util.spec_from_file_location("generate_nightly_perf_excel", module_path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _cell_value_by_header(ws, header_name: str, row_idx: int = 2):
    headers = [c.value for c in ws[1]]
    col_idx = headers.index(header_name) + 1
    return ws.cell(row=row_idx, column=col_idx).value


def test_generate_excel_report_with_perf_templates(tmp_path: Path):
    module = _load_excel_module()
    repo_root = Path(__file__).resolve().parents[1]
    perf_scripts_dir = repo_root / "tests" / "dfx" / "perf" / "scripts"

    omni_template_path = perf_scripts_dir / "result_omni_template.json"
    diffusion_template_path = perf_scripts_dir / "diffusion_result_template.json"

    omni_record = json.loads(omni_template_path.read_text(encoding="utf-8"))
    diffusion_records = json.loads(diffusion_template_path.read_text(encoding="utf-8"))

    input_dir = tmp_path / "input"
    diffusion_input_dir = tmp_path / "diffusion_input"
    input_dir.mkdir()
    diffusion_input_dir.mkdir()

    # Keep file names compatible with parser conventions in generate_nightly_perf_excel.py
    omni_result_file = input_dir / "result_test_perf_random_1_4_in2500_out900_20260415-185642.json"
    diffusion_result_file = diffusion_input_dir / "diffusion_result_qwen_image_edit_20260415-193200.json"
    omni_result_file.write_text(json.dumps(omni_record, ensure_ascii=False, indent=2), encoding="utf-8")
    diffusion_result_file.write_text(json.dumps(diffusion_records, ensure_ascii=False, indent=2), encoding="utf-8")

    output_file = tmp_path / "nightly_perf.xlsx"
    module.generate_excel_report(
        input_dir=str(input_dir),
        diffusion_input_dir=str(diffusion_input_dir),
        output_file=str(output_file),
        commit_sha="test_commit_sha",
        build_id="test_build_id",
        build_url="https://example.com/build/123",
    )

    assert output_file.exists()

    wb = load_workbook(output_file)
    assert set(wb.sheetnames) >= {"omni_summary", "diffusion_summary", "omni_raw", "diffusion_raw"}

    ws_omni_raw = wb["omni_raw"]
    baseline_cell = _cell_value_by_header(ws_omni_raw, "baseline")
    assert baseline_cell == json.dumps(omni_record["baseline"], ensure_ascii=False, sort_keys=True)

    ws_omni_summary = wb["omni_summary"]
    assert _cell_value_by_header(ws_omni_summary, "commit_sha") == "test_commit_sha"
    assert _cell_value_by_header(ws_omni_summary, "build_id") == "test_build_id"
