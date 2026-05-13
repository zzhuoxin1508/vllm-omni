# tests/test_omni_torch_profiler.py
from __future__ import annotations

import gzip
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

import vllm_omni.profiler.omni_torch_profiler as profiler_mod
from vllm_omni.profiler.omni_torch_profiler import OmniTorchProfilerWrapper


@pytest.fixture(autouse=True)
def patch_worker_profiler_init(monkeypatch):
    def fake_init(self, profiler_config):
        self.profiler_config = profiler_config

    monkeypatch.setattr(
        profiler_mod.WorkerProfiler,
        "__init__",
        fake_init,
    )


@dataclass
class DummyProfilerConfig:
    torch_profiler_dir: str
    torch_profiler_use_gzip: bool = False
    torch_profiler_record_shapes: bool = True
    torch_profiler_with_memory: bool = True
    torch_profiler_with_stack: bool = True
    torch_profiler_with_flops: bool = False
    torch_profiler_dump_cuda_time_total: bool = False


class FakeEvent:
    def __init__(
        self,
        *,
        name: str = "aten::mm",
        count: int = 1,
        input_shapes=None,
        stack=None,
        self_cpu_time_total: float = 10.0,
        cpu_time_total: float = 12.0,
        self_cuda_time_total: float = 20.0,
        cuda_time_total: float = 25.0,
        self_xpu_time_total: float = 0.0,
        xpu_time_total: float = 0.0,
        self_cpu_memory_usage: int = 128,
        cpu_memory_usage: int = 256,
        self_cuda_memory_usage: int = 1024,
        cuda_memory_usage: int = 2048,
        self_xpu_memory_usage: int = 0,
        xpu_memory_usage: int = 0,
        device_type: str = "CUDA",
        node_id: int = 0,
        overload_name: str = "",
        is_async: bool = False,
        is_legacy: bool = False,
    ):
        self.key = name
        self.name = name
        self.count = count
        self.input_shapes = input_shapes if input_shapes is not None else [[2, 2], [2, 2]]
        self.stack = stack if stack is not None else ["frame_a", "frame_b"]
        self.self_cpu_time_total = self_cpu_time_total
        self.cpu_time_total = cpu_time_total
        self.self_cuda_time_total = self_cuda_time_total
        self.cuda_time_total = cuda_time_total
        self.self_xpu_time_total = self_xpu_time_total
        self.xpu_time_total = xpu_time_total
        self.self_cpu_memory_usage = self_cpu_memory_usage
        self.cpu_memory_usage = cpu_memory_usage
        self.self_cuda_memory_usage = self_cuda_memory_usage
        self.cuda_memory_usage = cuda_memory_usage
        self.self_xpu_memory_usage = self_xpu_memory_usage
        self.xpu_memory_usage = xpu_memory_usage
        self.device_type = device_type
        self.node_id = node_id
        self.overload_name = overload_name
        self.is_async = is_async
        self.is_legacy = is_legacy


class FakeEventList(list):
    def table(self, sort_by=None, row_limit=-1):
        return f"fake_table(sort_by={sort_by}, row_limit={row_limit}, len={len(self)})"


class FakeTorchProfiler:
    def __init__(self, on_trace_ready=None):
        self.started = False
        self.stopped = False
        self.on_trace_ready = on_trace_ready
        self.exported_traces = []
        self.exported_stacks = []

    def start(self):
        self.started = True

    def stop(self):
        self.stopped = True
        if self.on_trace_ready is not None:
            self.on_trace_ready(self)

    def export_chrome_trace(self, path):
        Path(path).write_text('{"traceEvents": []}')
        self.exported_traces.append(path)

    def export_stacks(self, path, metric):
        Path(path).write_text(f"metric={metric}\nstack_line_1\nstack_line_2\n")
        self.exported_stacks.append((path, metric))

    def key_averages(self, group_by_input_shape=False, group_by_stack_n=0):
        if group_by_input_shape:
            return FakeEventList(
                [
                    FakeEvent(
                        name="aten::bmm",
                        input_shapes=[[4, 8, 16], [4, 16, 32]],
                    )
                ]
            )
        if group_by_stack_n:
            return FakeEventList(
                [
                    FakeEvent(
                        name="aten::all_reduce",
                        stack=["python_a", "python_b", "python_c"],
                    )
                ]
            )
        return FakeEventList(
            [
                FakeEvent(name="aten::mm"),
                FakeEvent(name="nccl:all_reduce"),
            ]
        )


@pytest.fixture
def fake_config(tmp_path):
    return DummyProfilerConfig(torch_profiler_dir=str(tmp_path))


@pytest.fixture
def fake_profiler_factory(monkeypatch):
    created = {}

    def fake_profile(*args, **kwargs):
        profiler = FakeTorchProfiler(on_trace_ready=kwargs.get("on_trace_ready"))
        created["profiler"] = profiler
        created["args"] = args
        created["kwargs"] = kwargs
        return profiler

    monkeypatch.setattr(profiler_mod.torch.profiler, "profile", fake_profile)
    return created


@pytest.fixture
def wrapper(fake_config, fake_profiler_factory):
    return OmniTorchProfilerWrapper(
        profiler_config=fake_config,
        worker_name="worker0",
        local_rank=0,
        activities=["CPU", "CUDA"],
    )


def test_set_trace_filename_creates_timestamped_session_dir(wrapper, monkeypatch, tmp_path):
    class FixedDatetime:
        @classmethod
        def now(cls):
            class _Now:
                def strftime(self, fmt):
                    return "20260403-034200"

            return _Now()

    monkeypatch.setattr(profiler_mod, "datetime", FixedDatetime)

    wrapper.set_trace_filename("stage_0_llm_1234567890")

    session_dir = Path(wrapper._session_dir)
    assert session_dir.exists()
    assert session_dir.parent == tmp_path
    assert session_dir.name == "20260403-034200_stage_0_llm_1234567890"


def test_set_trace_filename_with_full_path_creates_timestamped_leaf(wrapper, monkeypatch, tmp_path):
    class FixedDatetime:
        @classmethod
        def now(cls):
            class _Now:
                def strftime(self, fmt):
                    return "20260403-111111"

            return _Now()

    monkeypatch.setattr(profiler_mod, "datetime", FixedDatetime)

    target = tmp_path / "nested" / "stage_x"
    wrapper.set_trace_filename(str(target))

    session_dir = Path(wrapper._session_dir)
    assert session_dir.exists()
    assert session_dir.parent == target.parent
    assert session_dir.name == "20260403-111111_stage_x"


def test_on_trace_ready_exports_trace_json(wrapper):
    wrapper.set_trace_filename("case_trace")

    wrapper._on_trace_ready(wrapper.profiler)

    trace_path = Path(wrapper._trace_path)
    assert trace_path.exists()
    assert trace_path.name == "trace_rank0.json"
    assert trace_path.read_text() == '{"traceEvents": []}'


def test_on_trace_ready_exports_gzip_trace(fake_config, fake_profiler_factory, monkeypatch):
    fake_config.torch_profiler_use_gzip = True

    wrapper = OmniTorchProfilerWrapper(
        profiler_config=fake_config,
        worker_name="worker0",
        local_rank=0,
        activities=["CPU", "CUDA"],
    )
    wrapper.set_trace_filename("case_gzip")

    def fake_popen(cmd):
        assert cmd[:2] == ["gzip", "-f"]
        src = Path(cmd[2])
        gz_path = src.with_suffix(src.suffix + ".gz")
        gz_path.write_bytes(gzip.compress(src.read_bytes()))
        src.unlink()

        class DummyProc:
            pass

        return DummyProc()

    monkeypatch.setattr(profiler_mod.subprocess, "Popen", fake_popen)

    wrapper._on_trace_ready(wrapper.profiler)

    assert wrapper._trace_path.endswith(".json.gz")
    gz_path = Path(wrapper._trace_path)
    assert gz_path.exists()
    assert gzip.decompress(gz_path.read_bytes()) == b'{"traceEvents": []}'


def test_start_enables_memory_history(wrapper, monkeypatch):
    calls = []

    monkeypatch.setattr(profiler_mod.torch.cuda, "is_available", lambda: True)

    def fake_record_memory_history(*args, **kwargs):
        calls.append((args, kwargs))

    monkeypatch.setattr(
        profiler_mod.torch.cuda.memory,
        "_record_memory_history",
        fake_record_memory_history,
    )

    wrapper.set_trace_filename("case_memory_start")
    wrapper._start()

    assert wrapper.profiler.started is True
    assert wrapper._memory_history_enabled is True
    assert len(calls) == 1
    assert calls[0][1]["enabled"] == "all"
    assert calls[0][1]["context"] == "all"
    assert calls[0][1]["stacks"] == "python"
    assert calls[0][1]["max_entries"] == 100000
    assert calls[0][1]["clear_history"] is True


def test_start_skips_memory_history_when_memory_disabled(fake_config, fake_profiler_factory, monkeypatch):
    fake_config.torch_profiler_with_memory = False

    wrapper = OmniTorchProfilerWrapper(
        profiler_config=fake_config,
        worker_name="worker0",
        local_rank=0,
        activities=["CPU", "CUDA"],
    )

    called = {"n": 0}

    monkeypatch.setattr(profiler_mod.torch.cuda, "is_available", lambda: True)

    def fake_record_memory_history(*args, **kwargs):
        called["n"] += 1

    monkeypatch.setattr(
        profiler_mod.torch.cuda.memory,
        "_record_memory_history",
        fake_record_memory_history,
    )

    wrapper.set_trace_filename("case_skip_memory")
    wrapper._start()

    assert called["n"] == 0
    assert wrapper._memory_history_enabled is False


def test_try_dump_memory_snapshot_writes_pickle(wrapper, monkeypatch):
    wrapper.set_trace_filename("case_snapshot")
    wrapper._memory_history_enabled = True
    wrapper._memory_history_backend = "CUDA"
    wrapper._memory_history_module = profiler_mod.torch.cuda.memory

    disable_calls = []

    def fake_record_memory_history(*args, **kwargs):
        disable_calls.append((args, kwargs))

    def fake_dump_snapshot(path):
        Path(path).write_bytes(b"fake pickle bytes")

    monkeypatch.setattr(
        profiler_mod.torch.cuda.memory,
        "_record_memory_history",
        fake_record_memory_history,
    )
    monkeypatch.setattr(
        profiler_mod.torch.cuda.memory,
        "_dump_snapshot",
        fake_dump_snapshot,
    )

    wrapper._try_dump_memory_snapshot()

    snapshot = Path(wrapper._artifact_paths["memory_snapshot"])
    assert snapshot.exists()
    assert snapshot.name == "memory_snapshot_rank0.pickle"
    assert snapshot.read_bytes() == b"fake pickle bytes"
    assert wrapper._memory_history_enabled is False

    assert disable_calls[-1][1]["enabled"] is None


def test_stop_always_dumps_memory_snapshot_on_success_path(wrapper, monkeypatch):
    wrapper.set_trace_filename("case_stop")

    record_calls = []
    dump_calls = []

    monkeypatch.setattr(profiler_mod.torch.cuda, "is_available", lambda: True)

    def fake_record_memory_history(*args, **kwargs):
        record_calls.append((args, kwargs))

    def fake_dump_snapshot(path):
        dump_calls.append(path)
        Path(path).write_bytes(b"snapshot-bytes")

    monkeypatch.setattr(
        profiler_mod.torch.cuda.memory,
        "_record_memory_history",
        fake_record_memory_history,
    )
    monkeypatch.setattr(
        profiler_mod.torch.cuda.memory,
        "_dump_snapshot",
        fake_dump_snapshot,
    )

    wrapper._start()
    wrapper._stop()

    session_dir = Path(wrapper._session_dir)

    assert wrapper.profiler.started is True
    assert wrapper.profiler.stopped is True
    assert (session_dir / "memory_snapshot_rank0.pickle").exists()
    assert len(dump_calls) == 1
    assert record_calls[0][1]["enabled"] == "all"
    assert record_calls[-1][1]["enabled"] is None


def test_on_stop_hook_generates_stack_and_excel_artifacts(wrapper):
    wrapper.set_trace_filename("case_artifacts")
    wrapper._on_stop_hook()

    session_dir = Path(wrapper._session_dir)

    assert not (session_dir / "ops_summary_rank0.txt").exists()
    assert not (session_dir / "ops_by_shape_rank0.txt").exists()
    assert not (session_dir / "ops_by_stack_rank0.txt").exists()
    assert (session_dir / "stacks_cpu_rank0.txt").exists()
    assert (session_dir / "stacks_cuda_rank0.txt").exists()
    assert (session_dir / "ops_rank0.xlsx").exists()


def test_excel_contains_expected_sheets(wrapper):
    wrapper.set_trace_filename("case_excel")
    wrapper._on_stop_hook()

    xlsx_path = Path(wrapper._session_dir) / "ops_rank0.xlsx"
    wb = load_workbook(xlsx_path)

    assert "summary" in wb.sheetnames
    assert "by_shape" in wb.sheetnames
    assert "by_stack" in wb.sheetnames


def test_excel_summary_has_expected_columns(wrapper):
    wrapper.set_trace_filename("case_excel_columns")
    wrapper._on_stop_hook()

    xlsx_path = Path(wrapper._session_dir) / "ops_rank0.xlsx"
    wb = load_workbook(xlsx_path)
    ws = wb["summary"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "name" in headers
    assert "count" in headers
    assert "self_cpu_time_total_us" in headers
    assert "self_cuda_time_total_us" in headers
    assert "self_cpu_memory_usage_bytes" in headers
    assert "self_cuda_memory_usage_bytes" in headers
    assert "input_shapes" in headers
    assert "stack" in headers


def test_get_results_returns_all_artifact_paths(wrapper, monkeypatch):
    wrapper.set_trace_filename("case_results")

    monkeypatch.setattr(profiler_mod.torch.cuda, "is_available", lambda: True)
    monkeypatch.setattr(
        profiler_mod.torch.cuda.memory,
        "_record_memory_history",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        profiler_mod.torch.cuda.memory,
        "_dump_snapshot",
        lambda path: Path(path).write_bytes(b"snapshot"),
    )

    wrapper._start()
    wrapper._stop()

    results = wrapper.get_results()

    assert "trace" in results
    assert "table" in results
    assert "session_dir" in results
    assert "ops" in results
    assert "memory_snapshot" in results
    assert Path(results["session_dir"]).exists()
    assert Path(results["ops"]).exists()
    assert Path(results["table"]).exists()
    assert Path(results["table"]).name == "ops_rank0.xlsx"
    assert Path(results["memory_snapshot"]).exists()


def test_start_uses_xpu_memory_history_when_available(wrapper, monkeypatch):
    calls = []

    def fake_record_memory_history(*args, **kwargs):
        calls.append((args, kwargs))

    fake_memory_module = SimpleNamespace(
        _record_memory_history=fake_record_memory_history,
    )
    monkeypatch.setattr(
        wrapper,
        "_resolve_memory_history_backend",
        lambda: ("XPU", fake_memory_module),
    )

    wrapper.set_trace_filename("case_xpu_memory_start")
    wrapper._start()

    assert wrapper._memory_history_enabled is True
    assert wrapper._memory_history_backend == "XPU"
    assert wrapper._memory_history_module is fake_memory_module
    assert calls[0][1]["enabled"] == "all"


def test_start_uses_npu_memory_history_when_available(wrapper, monkeypatch):
    calls = []

    def fake_record_memory_history(*args, **kwargs):
        calls.append((args, kwargs))

    fake_memory_module = SimpleNamespace(
        _record_memory_history=fake_record_memory_history,
    )
    monkeypatch.setattr(
        wrapper,
        "_resolve_memory_history_backend",
        lambda: ("NPU", fake_memory_module),
    )

    wrapper.set_trace_filename("case_npu_memory_start")
    wrapper._start()

    assert wrapper._memory_history_enabled is True
    assert wrapper._memory_history_backend == "NPU"
    assert wrapper._memory_history_module is fake_memory_module
    assert calls[0][1]["enabled"] == "all"


def test_start_skips_memory_history_when_backend_api_missing(wrapper, monkeypatch):
    fake_memory_module = SimpleNamespace()
    monkeypatch.setattr(
        wrapper,
        "_resolve_memory_history_backend",
        lambda: ("XPU", fake_memory_module),
    )

    wrapper.set_trace_filename("case_missing_memory_api")
    wrapper._start()

    assert wrapper._memory_history_enabled is False
    assert wrapper._memory_history_backend is None
    assert wrapper._memory_history_module is None


def test_try_dump_memory_snapshot_uses_resolved_backend_module(wrapper):
    wrapper.set_trace_filename("case_xpu_snapshot")
    wrapper._memory_history_enabled = True
    wrapper._memory_history_backend = "XPU"

    calls = []

    def fake_record_memory_history(*args, **kwargs):
        calls.append((args, kwargs))

    def fake_dump_snapshot(path):
        Path(path).write_bytes(b"xpu snapshot bytes")

    wrapper._memory_history_module = SimpleNamespace(
        _record_memory_history=fake_record_memory_history,
        _dump_snapshot=fake_dump_snapshot,
    )

    wrapper._try_dump_memory_snapshot()

    snapshot = Path(wrapper._artifact_paths["memory_snapshot"])
    assert snapshot.exists()
    assert snapshot.read_bytes() == b"xpu snapshot bytes"
    assert calls[-1][1]["enabled"] is None
    assert wrapper._memory_history_enabled is False
    assert wrapper._memory_history_backend is None
    assert wrapper._memory_history_module is None


def test_event_list_to_rows_contains_expected_fields(wrapper):
    rows = wrapper._event_list_to_rows(
        [
            FakeEvent(
                name="aten::linear",
                input_shapes=[[8, 16], [16, 32]],
                stack=["f1", "f2"],
            )
        ]
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["name"] == "aten::linear"
    assert row["count"] == 1
    assert row["self_cpu_time_total_us"] == 10.0
    assert row["self_cuda_time_total_us"] == 20.0
    assert row["self_cpu_memory_usage_bytes"] == 128
    assert row["self_cuda_memory_usage_bytes"] == 1024
    assert "[[8, 16], [16, 32]]" == row["input_shapes"]
    assert row["stack"] == "f1\nf2"
